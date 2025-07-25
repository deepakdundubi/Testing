import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)


def get_data_for_Login(file, sheetName):
    final_list = []
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_column = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_column + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
    return final_list


def get_data_for_UMO(file, sheetName):
    final_list = []
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_column = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_column+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list[0])
    return final_list


def get_data_for_updateUMO(file, sheetName):
    final_list = []
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_column = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_column+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list[0])
    return final_list
